"""
load_ei_production.py
Extracts 2024 oil production and consumption (thousand barrels/day)
from EI Statistical Review Excel and loads into Neo4j Country nodes.

Run from project root:
    python scripts/load_ei_production.py
"""

import os
import openpyxl
from dotenv import load_dotenv
from neo4j import GraphDatabase

load_dotenv()

URI      = os.getenv("NEO4J_URI")
USER     = os.getenv("NEO4J_USERNAME")
PASSWORD = os.getenv("NEO4J_PASSWORD")

EXCEL_FILE = "data/raw/ei_stats_review_2025.xlsx"
TARGET_YEAR = 2024

# Map EI country names to ISO3
COUNTRY_NAME_TO_ISO3 = {
    "Algeria":              "DZA",
    "Angola":               "AGO",
    "Azerbaijan":           "AZE",
    "Belarus":              "BLR",
    "Djibouti":             "DJI",
    "Egypt":                "EGY",
    "Ethiopia":             "ETH",
    "Georgia":              "GEO",
    "Hungary":              "HUN",
    "Indonesia":            "IDN",
    "Iran":                 "IRN",
    "Iraq":                 "IRQ",
    "Kazakhstan":           "KAZ",
    "Kuwait":               "KWT",
    "Libya":                "LBY",
    "Malaysia":             "MYS",
    "Nigeria":              "NGA",
    "Oman":                 "OMN",
    "Pakistan":             "PAK",
    "Poland":               "POL",
    "Qatar":                "QAT",
    "Russia":               "RUS",
    "Saudi Arabia":         "SAU",
    "Singapore":            "SGP",
    "Somalia":              "SOM",
    "South Sudan":          "SSD",
    "Sudan":                "SDN",
    "Turkey":               "TUR",
    "Ukraine":              "UKR",
    "United Arab Emirates": "ARE",
    "Yemen":                "YEM",
}


def extract_sheet(wb, sheet_name, target_year):
    """Extract country → value for target_year from an EI sheet."""
    ws = wb[sheet_name]
    results = {}

    year_col = None
    header_row = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # Find header row with years
        if row[0] == "Thousand barrels daily":
            header_row = i
            for j, val in enumerate(row):
                if val == target_year:
                    year_col = j
                    break
            break

    if year_col is None:
        print(f"  WARNING: could not find year {target_year} in {sheet_name}")
        return results

    # Now read data rows
    in_data = False
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row:
            continue

        country_name = row[0]
        if not isinstance(country_name, str):
            continue

        # Stop at totals/regional aggregates
        if any(x in country_name for x in ["Total", "Other", "of which", "World"]):
            continue

        iso3 = COUNTRY_NAME_TO_ISO3.get(country_name.strip())
        if not iso3:
            continue

        val = row[year_col]
        if isinstance(val, (int, float)):
            results[iso3] = round(float(val), 2)

    return results


def main():
    print(f"Loading {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    print(f"Extracting production data for {TARGET_YEAR}...")
    production = extract_sheet(wb, "Oil Production - barrels", TARGET_YEAR)
    print(f"  Found {len(production)} countries")

    print(f"Extracting consumption data for {TARGET_YEAR}...")
    consumption = extract_sheet(wb, "Oil Consumption - barrels", TARGET_YEAR)
    print(f"  Found {len(consumption)} countries")

    # Merge into one dict
    all_countries = set(list(production.keys()) + list(consumption.keys()))
    print(f"\nLoading into Neo4j ({len(all_countries)} countries)...")

    driver = GraphDatabase.driver(URI, auth=(USER, PASSWORD))
    loaded = 0

    with driver.session() as session:
        for iso3 in all_countries:
            prod = production.get(iso3)
            cons = consumption.get(iso3)

            if prod is not None:
                session.run("""
                    MATCH (c:Country {iso3: $iso3})
                    SET c.production_kbpd = $val,
                        c.production_source = 'EI Statistical Review 2025'
                """, iso3=iso3, val=prod)

            if cons is not None:
                session.run("""
                    MATCH (c:Country {iso3: $iso3})
                    SET c.consumption_kbpd = $val,
                        c.consumption_source = 'EI Statistical Review 2025'
                """, iso3=iso3, val=cons)

            loaded += 1

    driver.close()

    print(f"Done. Updated {loaded} countries.")
    print("\nSample production (thousand bpd):")
    for iso3, val in list(production.items())[:5]:
        print(f"  {iso3}: {val}")
    print("\nSample consumption (thousand bpd):")
    for iso3, val in list(consumption.items())[:5]:
        print(f"  {iso3}: {val}")


if __name__ == "__main__":
    main()
